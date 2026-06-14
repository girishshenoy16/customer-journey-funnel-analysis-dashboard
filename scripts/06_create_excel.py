import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv("data/customer_journey_cleaned.csv")

wb = Workbook()

header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
sub_header_fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=14, color='1A365D')
kpi_font = Font(name='Calibri', bold=True, size=28, color='1A365D')
kpi_label_font = Font(name='Calibri', size=11, color='666666')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def auto_width(ws, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[chr(64 + col)].width = min(max_len + 3, 30)

ws1 = wb.active
ws1.title = "Dashboard Summary"
ws1.merge_cells('A1:F1')
ws1['A1'] = "CUSTOMER JOURNEY & FUNNEL ANALYSIS - DASHBOARD SUMMARY"
ws1['A1'].font = Font(name='Calibri', bold=True, size=16, color='1A365D')
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A3:C3')
ws1['A3'] = "KEY PERFORMANCE INDICATORS"
ws1['A3'].font = title_font

kpis = [
    ('Total Visitors', len(df)),
    ('Leads Generated', df['Product Viewed'].sum()),
    ('Add to Cart', df['Add to Cart'].sum()),
    ('Checkout Started', df['Checkout Started'].sum()),
    ('Purchase Completed', df['Purchase Completed'].sum()),
    ('Conversion Rate (%)', round(df['Purchase Completed'].sum() / len(df) * 100, 1)),
    ('Retention Rate (%)', round(len(df[df['Customer Retention Status'] == 'Retained']) / len(df) * 100, 1)),
    ('Total Revenue ($)', round(df['Purchase Value'].sum(), 2)),
    ('Avg Purchase Value ($)', round(df[df['Purchase Completed'] == 1]['Purchase Value'].mean(), 2)),
    ('Avg Feedback Score', round(df['Customer Feedback Score'].mean(), 2))
]

ws1['A5'] = 'Metric'
ws1['B5'] = 'Value'
style_header_row(ws1, 5, 2)

for i, (metric, value) in enumerate(kpis, start=6):
    ws1.cell(row=i, column=1, value=metric).font = Font(bold=True, size=11)
    ws1.cell(row=i, column=2, value=value).alignment = Alignment(horizontal='center')
    ws1.cell(row=i, column=2).font = Font(size=11)
    ws1.cell(row=i, column=1).border = thin_border
    ws1.cell(row=i, column=2).border = thin_border

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 15

ws2 = wb.create_sheet("Funnel Analysis")
ws2.merge_cells('A1:D1')
ws2['A1'] = "FUNNEL ANALYSIS"
ws2['A1'].font = title_font

funnel_data = [
    ('Stage', 'Count', '% of Visitors', 'Drop-off %'),
    ('Website Visit', len(df), 100.0, '-'),
    ('Product Viewed', df['Product Viewed'].sum(), 100.0, 0.0),
    ('Add to Cart', df['Add to Cart'].sum(), round(df['Add to Cart'].sum() / len(df) * 100, 1),
     round((df['Product Viewed'].sum() - df['Add to Cart'].sum()) / df['Product Viewed'].sum() * 100, 1)),
    ('Checkout Started', df['Checkout Started'].sum(), round(df['Checkout Started'].sum() / len(df) * 100, 1),
     round((df['Add to Cart'].sum() - df['Checkout Started'].sum()) / df['Add to Cart'].sum() * 100, 1)),
    ('Purchase Completed', df['Purchase Completed'].sum(), round(df['Purchase Completed'].sum() / len(df) * 100, 1),
     round((df['Checkout Started'].sum() - df['Purchase Completed'].sum()) / df['Checkout Started'].sum() * 100, 1))
]

for i, row_data in enumerate(funnel_data, start=3):
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=i, column=j + 1, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if i == 3:
            cell.font = header_font
            cell.fill = header_fill

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15

ws3 = wb.create_sheet("Channel Performance")
ws3.merge_cells('A1:G1')
ws3['A1'] = "CHANNEL PERFORMANCE ANALYSIS"
ws3['A1'].font = title_font

channel_headers = ['Channel', 'Visitors', 'Conversions', 'Conv Rate %', 'Revenue ($)', 'Retention %', 'Est. CAC ($)']
for j, h in enumerate(channel_headers, start=1):
    ws3.cell(row=3, column=j, value=h).font = header_font
    ws3.cell(row=3, column=j).fill = header_fill
    ws3.cell(row=3, column=j).alignment = Alignment(horizontal='center')
    ws3.cell(row=3, column=j).border = thin_border

# Fixed Cost Per Lead (CPL) per channel to ensure consistent, non-random calculations
CPL_MAP = {
    'Paid Ads': 12.50,
    'Social Media': 8.20,
    'Organic Search': 1.50,
    'Email': 2.10,
    'Referral': 3.50,
    'Direct': 0.50
}

channels = df['Acquisition Channel'].unique()
row_idx = 4
for ch in sorted(channels):
    subset = df[df['Acquisition Channel'] == ch]
    total = len(subset)
    conv = subset['Purchase Completed'].sum()
    rev = round(subset['Purchase Value'].sum(), 2)
    ret = round(len(subset[subset['Customer Retention Status'] == 'Retained']) / total * 100, 1)
    conv_rate = round(conv / total * 100, 1)
    cost_per_lead = CPL_MAP.get(ch, 5.00)
    cac = round(cost_per_lead / (conv / total) if conv > 0 else 0, 2)
    values = [ch, total, conv, conv_rate, rev, ret, cac]
    for j, v in enumerate(values, start=1):
        cell = ws3.cell(row=row_idx, column=j, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

for col in range(1, 8):
    ws3.column_dimensions[chr(64 + col)].width = 18

ws4 = wb.create_sheet("Segment Analysis")
ws4.merge_cells('A1:F1')
ws4['A1'] = "SEGMENT-WISE CONVERSION ANALYSIS"
ws4['A1'].font = title_font

seg_headers = ['Segment', 'Total', 'Converted', 'Conv Rate %', 'Retention %', 'Avg Revenue ($)']
for j, h in enumerate(seg_headers, start=1):
    ws4.cell(row=3, column=j, value=h).font = header_font
    ws4.cell(row=3, column=j).fill = header_fill
    ws4.cell(row=3, column=j).alignment = Alignment(horizontal='center')
    ws4.cell(row=3, column=j).border = thin_border

segments = ['Premium', 'Standard', 'Budget']
row_idx = 4
for seg in segments:
    subset = df[df['Customer Segment'] == seg]
    total = len(subset)
    conv = subset['Purchase Completed'].sum()
    ret = round(len(subset[subset['Customer Retention Status'] == 'Retained']) / total * 100, 1)
    conv_rate = round(conv / total * 100, 1)
    avg_rev = round(subset[subset['Purchase Completed'] == 1]['Purchase Value'].mean() if conv > 0 else 0, 2)
    values = [seg, total, conv, conv_rate, ret, avg_rev]
    for j, v in enumerate(values, start=1):
        cell = ws4.cell(row=row_idx, column=j, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

for col in range(1, 7):
    ws4.column_dimensions[chr(64 + col)].width = 18

ws5 = wb.create_sheet("Regional Analysis")
ws5.merge_cells('A1:F1')
ws5['A1'] = "REGIONAL PERFORMANCE ANALYSIS"
ws5['A1'].font = title_font

reg_headers = ['Region', 'Total', 'Converted', 'Conv Rate %', 'Retention %', 'Avg Revenue ($)']
for j, h in enumerate(reg_headers, start=1):
    ws5.cell(row=3, column=j, value=h).font = header_font
    ws5.cell(row=3, column=j).fill = header_fill
    ws5.cell(row=3, column=j).alignment = Alignment(horizontal='center')
    ws5.cell(row=3, column=j).border = thin_border

regions = ['North', 'South', 'East', 'West', 'Central']
row_idx = 4
for reg in regions:
    subset = df[df['Region'] == reg]
    total = len(subset)
    conv = subset['Purchase Completed'].sum()
    ret = round(len(subset[subset['Customer Retention Status'] == 'Retained']) / total * 100, 1)
    conv_rate = round(conv / total * 100, 1)
    avg_rev = round(subset[subset['Purchase Completed'] == 1]['Purchase Value'].mean() if conv > 0 else 0, 2)
    values = [reg, total, conv, conv_rate, ret, avg_rev]
    for j, v in enumerate(values, start=1):
        cell = ws5.cell(row=row_idx, column=j, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

for col in range(1, 7):
    ws5.column_dimensions[chr(64 + col)].width = 18

ws6 = wb.create_sheet("Campaign Analysis")
ws6.merge_cells('A1:G1')
ws6['A1'] = "CAMPAIGN PERFORMANCE ANALYSIS"
ws6['A1'].font = title_font

camp_headers = ['Campaign', 'Visitors', 'Conversions', 'Conv Rate %', 'Revenue ($)', 'Retention %', 'Avg Feedback']
for j, h in enumerate(camp_headers, start=1):
    ws6.cell(row=3, column=j, value=h).font = header_font
    ws6.cell(row=3, column=j).fill = header_fill
    ws6.cell(row=3, column=j).alignment = Alignment(horizontal='center')
    ws6.cell(row=3, column=j).border = thin_border

campaigns = df['Campaign Name'].unique()
row_idx = 4
for camp in sorted(campaigns):
    subset = df[df['Campaign Name'] == camp]
    total = len(subset)
    conv = subset['Purchase Completed'].sum()
    rev = round(subset['Purchase Value'].sum(), 2)
    ret = round(len(subset[subset['Customer Retention Status'] == 'Retained']) / total * 100, 1)
    conv_rate = round(conv / total * 100, 1)
    avg_fb = round(subset['Customer Feedback Score'].mean(), 2)
    values = [camp, total, conv, conv_rate, rev, ret, avg_fb]
    for j, v in enumerate(values, start=1):
        cell = ws6.cell(row=row_idx, column=j, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

for col in range(1, 8):
    ws6.column_dimensions[chr(64 + col)].width = 20

ws7 = wb.create_sheet("Retention Analysis")
ws7.merge_cells('A1:G1')
ws7['A1'] = "CUSTOMER RETENTION ANALYSIS"
ws7['A1'].font = title_font

ret_headers = ['Customer Segment', 'Total', 'Retained', 'Churned', 'At Risk', 'Retention %', 'Avg Feedback']
for j, h in enumerate(ret_headers, start=1):
    ws7.cell(row=3, column=j, value=h).font = header_font
    ws7.cell(row=3, column=j).fill = header_fill
    ws7.cell(row=3, column=j).alignment = Alignment(horizontal='center')
    ws7.cell(row=3, column=j).border = thin_border

row_idx = 4
for seg in segments:
    subset = df[df['Customer Segment'] == seg]
    total = len(subset)
    retained = len(subset[subset['Customer Retention Status'] == 'Retained'])
    churned = len(subset[subset['Customer Retention Status'] == 'Churned'])
    at_risk = len(subset[subset['Customer Retention Status'] == 'At Risk'])
    ret_rate = round(retained / total * 100, 1)
    avg_fb = round(subset['Customer Feedback Score'].mean(), 2)
    values = [seg, total, retained, churned, at_risk, ret_rate, avg_fb]
    for j, v in enumerate(values, start=1):
        cell = ws7.cell(row=row_idx, column=j, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

for col in range(1, 8):
    ws7.column_dimensions[chr(64 + col)].width = 20

wb.save("excel/customer_journey_analysis.xlsx")
print("Excel workbook saved to excel/customer_journey_analysis.xlsx")
print(f"Sheets: {wb.sheetnames}")
